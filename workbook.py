# import workbook as workbook
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors, Font, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment




# # 加载文件
# wb = load_workbook('./5a.xlsx')
#
# # 读取 sheetname
# print('输入文件所有工作表名：\n', wb.sheetnames)
# ws = wb['5a']
#
# # 或者不知道名字时
# sheet_names = wb.sheetnames
# ws2 = wb[sheet_names[0]]    # index 为 0 为第一张表
# print(ws is ws2)
#
# ws.title = '_5a'
# print('修改sheetname: \n', wb.sheetnames)

workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "hello"
sheet["B1"] = "world!"

#workbook.save(filename="hello_world.xlsx")

workbook = load_workbook(filename="5a.xlsx")  # 加载文件
workbook.sheetnames                          # 查看所有的工作薄

print(workbook.sheetnames)

wb = workbook.sheetnames[0]               # 选择第一个工作薄
print("wb", wb)


sheet = workbook.active                     # 选择最后一个工作薄
print(sheet)

sheet.title                                # 选择的工作薄
print("title", sheet.title)




b = sheet["A1"]                       # 定位工作薄 A1 列位置
print("A1+ ",b)

a = sheet["A1"].value                 # 读取 A1 列的值
c = sheet["C1"].value                 # 读取 C1 列的值
print(a)
print(c)

d = sheet.cell(row=2,column=2)         # 定位工作薄 2 行 2 列
d1 = sheet.cell(row=2,column=2).value  # 读取 2 行 2 列
print("d", d)
print("d1", d1)



e = sheet["A1":"B2"]               # 定义工作表的区域块，如 A1 B1 A2 B2
# e1 = sheet["A1":"B2"].value
print("e", e)
# print("e1", e1)



f = sheet["A"]      # 定位 A 的这一整列
print("f", f)

g = sheet[3]       # 定位 第三行
g1 = sheet[2:5]       # 定位 第二，三，四，五行
print("g", g)
print("g1", g1)

# 使用 for 循环来定义位置区域块，以行的形式读取区域块坐标
for row in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=3):
    print("row", row)
# 使用 for 循环来定义位置区域块，以列的形式读取区域块坐标
for column in sheet.iter_cols(min_row=1,max_row=2,min_col=1,max_col=3):
    print(column)
# 以行读取坐标块，打印坐标块的值
for value in sheet.iter_rows(min_row=1,
                             max_row=2,
                             min_col=1,
                             max_col=3,
                             values_only=True):
    print(value)


workbook = load_workbook("sample.xlsx")

sheet = workbook.active
products = {}

for row1 in sheet.iter_rows(min_row=2,
                            min_col=1,
                            max_col=5,
                            values_only=True):
    print(row1)
